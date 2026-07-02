from flask import Flask, render_template, g, request, redirect, url_for, flash, get_flashed_messages,send_file
from werkzeug.utils import secure_filename
import os.path
from pathlib import Path
import shutil
from datetime import datetime
from openpyxl import Workbook, worksheet
from io import BytesIO
import json
import zipfile

import detection
from database import try_connect,DATABASE,select_data,insert_data,clear_data,select_by_id

UPLOAD_FOLDER = "./detections/"
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
header_links = [{"name": "Главная страница", "url": "index"},
                {"name": "Обработка изображений", "url": "process"},
                {"name": "Представление данных", "url": "dataview"}]

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


def get_db():
    """Создание соединения с БД"""

    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = try_connect(DATABASE)
    return db


@app.teardown_appcontext
def close_connection(exception):
    """Закрытие соединения с БД перед остановкой сервера"""
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


@app.route('/index')
@app.route('/')
def index():
    """Переход на главвную страницу веб-приложения"""
    return render_template('index.html', header_links=header_links)


@app.route('/dataview')
def dataview():
    """Переход на страницу "Представление данных\""""
    db = get_db()
    records = select_data(db)
    for rec in records:
        rec['datetime_detect'] = datetime.fromtimestamp(rec['datetime_detect']).strftime('%d.%m.%Y %H:%M:%S')
    return render_template('dataview.html', header_links=header_links, data=records)


@app.route('/process')
def process():
    """Переход на страницу "Обработка изображений\""""
    return render_template('process.html', header_links=header_links)


@app.route('/download',methods=["POST", "GET"])
def download():
    """Скачивание файла по "url\""""
    url = request.args.get('url')
    return send_file(url,as_attachment=True)

@app.route('/download_zip',methods=["POST", "GET"])
def download_zip():
    """Скачивание ZIP-файла с результатами обработки"""
    id_file= request.args.get('id')
    db = get_db()
    record = dict(select_by_id(db,int(id_file)))
    zipfile_bin = BytesIO()

    with zipfile.ZipFile(zipfile_bin,'w',zipfile.ZIP_DEFLATED) as zip_f:
        zip_f.write( record['source_img_url'],
                     arcname=os.path.join("/results/", record['source_img_url'].split('/')[-1]) )
        zip_f.write(record['result_img_url'],
                    arcname=os.path.join("/results/", record['result_img_url'].split('/')[-1]))
        zip_f.write(record['json_data_url'],
                    arcname=os.path.join("/results/", record['json_data_url'].split('/')[-1]))

    zipfile_bin.seek(0)

    filename = record['source_img_url'].split('/')[-1].split('.')[0]
    return send_file(zipfile_bin,as_attachment=True,download_name=filename+'.zip',mimetype="application/zip")


@app.route('/make_report',methods=["POST", "GET"])
def make_report():
    """Формирование отчета по истории обработок из БД"""
    db = get_db()
    records = select_data(db)
    for rec in records:
        rec['datetime_detect'] = datetime.fromtimestamp(rec['datetime_detect']).strftime('%d.%m.%Y %H:%M:%S')

    wb: Workbook = Workbook()
    ws: worksheet = wb.active
    ws.append(["Дата, время", "Имя файла", "Количество обнаруженных судов"])
    for row, rec in enumerate(records):
        ws.cell(row=row + 2, column=1, value=rec['datetime_detect'])
        ws.cell(row=row + 2, column=2, value=rec['source_img_url'].split('/')[-1])
        with open(rec['json_data_url'], 'r') as f_json:
            json_data = json.load(f_json)
        ws.cell(row=row + 2, column=3, value=json_data['count'])

    bin_file = BytesIO()
    wb.save(bin_file)
    bin_file.seek(0)

    return send_file(bin_file, as_attachment=True, download_name='Отчет.xlsx',
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/clear_db',methods=["POST"])
def clear_db():
    """Очистка истории БД"""
    db = get_db()

    result = clear_data(db)
    if result:
        flash("История очищена.", category="good")
    else:
        flash("Не удалось очистить историю.", category="bad")

    dirs = os.listdir(UPLOAD_FOLDER)
    for detect_dir in dirs:
        shutil.rmtree(os.path.join(UPLOAD_FOLDER,detect_dir))

    return redirect(url_for('dataview'))


@app.errorhandler(413)
def request_entity_too_large(error):
    """Перехватчик исключения в случае загрузки файла размером большего допустимого"""
    flash("Размер файла превышает допустимый размер (16 МБ). Выберите другой файл.", category="bad")
    return redirect(url_for('process'))


@app.route('/upload_img', methods=["POST", "GET"])
def upload_img():
    """Загрузка фото для обработки"""
    if request.method == 'POST':
        files = request.files.getlist('img')
        for file in files:

            if file.filename != '' and file.filename.split('.')[-1] in ALLOWED_EXTENSIONS:
                # Получение названия файла
                filename = secure_filename(file.filename)

                # Получение текущей даты и времени
                current_date_time = datetime.now()
                str_date_time = current_date_time.strftime('%Y_%m_%d_%H_%M_%S_%f')

                # Определение и создание директории с результатми обработки
                Path(app.config['UPLOAD_FOLDER']).mkdir(parents=True, exist_ok=True)

                file_dir = os.path.join(app.config['UPLOAD_FOLDER'], "{0}/".format(str_date_time))
                Path(file_dir).mkdir(parents=True, exist_ok=True)

                # Сохранение файла на обработку
                file_path = os.path.join(file_dir, filename)
                file.save(file_path)


                # Запуск обработки
                detection_result = detection.process_img(file_dir, filename)

                # Если обработка успешна
                if detection_result:
                    flash(f"Обработка {filename} выполнена.", category="good")
                    result_name = "result_" + filename
                    json_name = "data.json"

                    record_data={"datetime_detect":current_date_time.timestamp(),
                                 "source_img_url":file_path,
                                 "result_img_url":os.path.join(file_dir, result_name),
                                 "json_data_url":os.path.join(file_dir, json_name)}

                    db = get_db()

                    # Добавление записи в БД
                    result_insert = insert_data(db, record_data)

                    # Если добавление в БД успешно
                    if result_insert:
                        flash("Результаты обработки сохранены.", category="good")
                    else:
                        shutil.rmtree(file_dir)
                        flash("Не удалось сохранить результаты обработки. Повторите снова.", category="bad")
                else:
                    shutil.rmtree(file_dir)
                    flash("Возникла ошибка во время обработки. Повторите обработку.", category="bad")

            else:
                files.pop()
                if len(files)==0:
                    flash(f"Не были выбраны файлы", category="bad")
                else:
                    flash(f"Выбран неккоректный файл: {file}", category="bad")

    return redirect(url_for('process'))



